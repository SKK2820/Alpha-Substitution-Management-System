import math
from datetime import date
from mysql.connector import IntegrityError
from datetime import datetime
from flask import (Flask,render_template,request,send_file,redirect,url_for,flash,session)
from database import get_absences_by_date, fetch_only_today_substitutions,search_upcoming_substitutions,fetch_upcoming_substitutions,search_previous_substitutions,search_upcoming_absences,fetch_upcoming_absences,search_previous_absences,update_teacher_full,get_teacher_name,get_dashboard_stats,generate_substitution_candidates,search_today_substitutions_with_absent_teacher,insert_substitutions_batch, get_teacher_classes ,get_all_absences_report,get_all_substitutions_report, update_user_password, update_admin_password, update_staff_password, get_staff_password, get_today_substitutions ,get_my_absences, get_my_substitutions, get_teacher_full_timetable, get_teacher_details, login_user ,create_database,initialize_database ,add_teacher, search_substitutions ,search_teachers,search_absences ,count_teachers,count_absences_today,count_substitutions_today_dashboard, count_unavailable_substitutions ,count_absent_teachers,fetch_substitutions,absence_exists, delete_absence, fetch_absences ,record_absence , get_teacher_id, fetch_all_teachers,update_timetable,get_full_timetable,delete_teacher, get_teacher_timetable ,get_teacher_names,update_teacher,get_teacher
from reportlab.platypus import (SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from openpyxl.styles import Font, Alignment,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from io import BytesIO



create_database()
initialize_database()

app = Flask(__name__)
app.secret_key = "alpha_substitution_management_system"

@app.route("/home")
def home():

    # Allow only admin
    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if session["role"] != "admin":

        return redirect(
            url_for(
                "teacher_dashboard"
            )
        )

    today = datetime.today()

    date = today.strftime(
        "%Y-%m-%d"
    )

    display_date = today.strftime(
        "%d %B %Y (%A)"
    )


    # Load all dashboard statistics
    # using ONE database connection

    stats = get_dashboard_stats(date)


    return render_template(

        "index.html",

        today=display_date,

        total_teachers=
        stats["total_teachers"],

        absences_today=
        stats["absences_today"],

        substitutions_today=
        stats["substitutions_today"],

        unavailable_today=
        stats["unavailable_today"]
    )

@app.route("/suggest_substitutions")
def suggest_substitutions():

    absence_ids = session.get(
        "pending_absence_ids",
        []
    )

    suggestions = generate_substitution_candidates(
        absence_ids
    )

    return render_template(

        "suggest_substitutions.html",

        suggestions=suggestions,
    )

@app.route(
    "/preview_substitutions",
    methods=["POST"])
def preview_substitutions():

    count = int(
        request.form["count"]
    )

    preview_data = []

    # Stores:
    # (normalized teacher name, period)
    assigned_substitutes = set()


    for i in range(1, count + 1):

        absence_id = int(
            request.form[f"absence_id_{i}"]
        )

        period = int(
            request.form[f"period_{i}"]
        )

        class_name = request.form[
            f"class_{i}"
        ]

        teacher_value = request.form[
            f"teacher_{i}"
        ]

        manual_substitute_name = None

        duplicate_key = None


        # Case 1:
        # Substitute not available
        if teacher_value == "NONE":

            teacher_id = None

            teacher_name = (
                "Substitute Not Available"
            )


        # Case 2:
        # Manual entry
        elif teacher_value == "MANUAL":

            teacher_id = None

            manual_substitute_name = request.form[
                f"manual_teacher_{i}"
            ].strip()


            if manual_substitute_name:

                teacher_name = (
                    manual_substitute_name
                    +
                    " (Manual Entry)"
                )

                # Normalize manual teacher name
                normalized_teacher_name = (
                    manual_substitute_name
                    .strip()
                    .casefold()
                )

                duplicate_key = (
                    normalized_teacher_name,
                    period
                )

            else:

                teacher_name = (
                    "Substitute Not Available"
                )


        # Case 3:
        # Normal teacher
        else:

            teacher_id = int(
                teacher_value
            )

            teacher_name = get_teacher_name(
                teacher_id
            )

            # Normalize database teacher name
            normalized_teacher_name = (
                teacher_name
                .strip()
                .casefold()
            )

            duplicate_key = (
                normalized_teacher_name,
                period
            )


        # Check duplicate assignment

        if duplicate_key is not None:

            if duplicate_key in assigned_substitutes:

                # Remove Manual Entry text
                # from warning message
                warning_teacher_name = (
                    manual_substitute_name
                    if manual_substitute_name
                    else teacher_name
                )

                flash(

                    f"{warning_teacher_name} is already "
                    f"assigned to another class during "
                    f"Period {period}. Please select a "
                    f"different substitute.",

                    "danger"
                )

                return redirect(
                    url_for(
                        "suggest_substitutions"
                    )
                )


            assigned_substitutes.add(
                duplicate_key
            )


        preview_data.append({

            "absence_id":
            absence_id,

            "period":
            period,

            "class_name":
            class_name,

            "teacher_id":
            teacher_id,

            "teacher_name":
            teacher_name,

            "manual_substitute_name":
            manual_substitute_name,

            "absent_teacher":
            request.form[
                f"absent_teacher_{i}"
            ]
        })


    session[
        "preview_substitutions"
    ] = preview_data


    return render_template(

        "preview_substitutions.html",

        substitutions=preview_data
    )

@app.route("/finalize_substitutions",methods=["POST"])
def finalize_substitutions():

    substitutions = session.get(
        "preview_substitutions",
        []
    )

    if substitutions:

        insert_substitutions_batch(
            substitutions
        )

    session.pop(
        "preview_substitutions",
        None
    )

    session.pop(
        "pending_absence_ids",
        None
    )

    flash(
        "Substitution timetable finalized successfully!",
        "success"
    )

    return redirect(
        url_for(
            "substitutions_page"
        )
    )

@app.route("/", methods=["GET", "POST"])
def login_page():

    # Already logged in?
    if "role" in session:

        if session["role"] == "admin":

            return redirect(
                url_for(
                    "home"
                )
            )

        else:

            return redirect(
                url_for(
                    "teacher_dashboard"
                )
            )

    if request.method == "POST":

        username = request.form["username"]

        password = request.form["password"]

        user = login_user(
            username,
            password
        )

        if user:

            session["user_id"] = user[0]

            session["username"] = user[1]

            session["role"] = user[2]

            session["teacher_id"] = user[3]

            if user[2] == "admin":

                return redirect(
                    url_for(
                        "home"
                    )
                )

            else:

                return redirect(
                    url_for(
                        "teacher_dashboard"
                    )
                )

        flash(
            "Invalid username or password!",
            "danger"
        )

    return render_template(
        "login.html"
    )

@app.route(
    "/change_admin_password",
    methods=["GET", "POST"]
)
def change_admin_password():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if session["role"] != "admin":

        return redirect(
            url_for(
                "teacher_dashboard"
            )
        )

    if request.method == "POST":

        current_password = request.form[
            "current_password"
        ]

        new_password = request.form[
            "new_password"
        ]

        confirm_password = request.form[
            "confirm_password"
        ]

        user = login_user(
            "admin",
            current_password
        )

        if user is None:

            flash(
                "Current password is incorrect!",
                "danger"
            )

        elif new_password != confirm_password:

            flash(
                "Passwords do not match!",
                "danger"
            )

        else:

            update_admin_password(
                new_password
            )

            flash(
                "Admin password changed successfully!",
                "success"
            )

            return redirect(
                url_for(
                    "home"
                )
            )

    return render_template(
        "change_password.html"
    )

@app.route("/logout")
def logout():

    session.clear()

    flash(
        "Logged out successfully!",
        "success"
    )

    return redirect(
        url_for(
            "login_page"
        )
    )

@app.route("/export_substitutions_pdf")
def export_substitutions_pdf():

    # ---------------------------------
    # GET TODAY'S DATE
    # ---------------------------------

    today = date.today()


    # ---------------------------------
    # FETCH ONLY TODAY'S SUBSTITUTIONS
    # ---------------------------------
    records = fetch_only_today_substitutions(today)

    # ---------------------------------
    # CREATE PDF
    # ---------------------------------

    buffer = BytesIO()

    pdf = SimpleDocTemplate(
        buffer,
    )


    # ---------------------------------
    # PDF STYLES
    # ---------------------------------

    styles = getSampleStyleSheet()


    school_heading = Paragraph(

        "<b>ALPHA SCHOOL, CIT NAGAR</b>",

        styles["Title"]
    )

    system_heading_style=styles["Heading2"].clone("SystemHeading")

    system_heading = Paragraph(

        "<b>SUBSTITUTION MANAGEMENT SYSTEM</b>",

        system_heading_style
    )


    formatted_date = today.strftime(
        "%d-%m-%Y"
    )


    date_heading = Paragraph(

        f"<b>Date: {formatted_date}</b>",

        styles["Heading3"]
    )


    # ---------------------------------
    # TABLE DATA
    # ---------------------------------

    data = [

        [
            "Period",
            "Absent Teacher",
            "Class",
            "Substitute Teacher"
        ]

    ]


    for row in records:

        period = row[0]

        absent_teacher = row[1]

        class_name = row[2]

        substitute_teacher = row[3]

        manual_substitute_name = row[4]


        # Normal database teacher
        if substitute_teacher:

            teacher_name = substitute_teacher


        # Manual substitute
        elif manual_substitute_name:

            teacher_name = manual_substitute_name


        # No substitute available
        else:

            teacher_name = (
                "Substitute Not Available"
            )


        data.append(

            [
                period,
                absent_teacher,
                class_name,
                teacher_name
            ]

        )


    # ---------------------------------
    # CREATE TABLE
    # ---------------------------------

    table = Table(
        data
    )


    table.setStyle(

        TableStyle(

            [

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.grey
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.whitesmoke
                ),

                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, -1),
                    colors.beige
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    1,
                    colors.black
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                )

            ]

        )

    )


    # ---------------------------------
    # BUILD PDF
    # ---------------------------------

    elements = [

        school_heading,

        Spacer(1, 8),

        system_heading,

        Spacer(1, 20),

        date_heading,

        Spacer(1, 15),

        table

    ]


    pdf.build(
        elements
    )


    buffer.seek(0)


    return send_file(

        buffer,

        as_attachment=True,

        download_name=(
            f"substitutions_{formatted_date}.pdf"
        ),

        mimetype="application/pdf"
    )

@app.route("/export_substitutions_excel")
def export_substitutions_excel():

    today = datetime.today()

    date = today.strftime(
        "%Y-%m-%d"
    )

    day = today.strftime(
        "%A"
    )

    records = get_today_substitutions(
        date
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "Substitutions"

    # Title
    ws.merge_cells(
        "A1:C1"
    )

    title = ws["A1"]

    title.value = (
    "ALPHA SUBSTITUTION MANAGEMENT SYSTEM"
)

    title.font = Font(
    bold=True,
    size=16
)

    title.alignment = Alignment(
    horizontal="center"
)

    # Subtitle
    ws.merge_cells(
    "A2:C2"
)

    subtitle = ws["A2"]

    subtitle.value = (
    f"Substitution Timetable for {date} ({day})"
)

    subtitle.font = Font(
    bold=True,
    size=13
)

    subtitle.alignment = Alignment(
    horizontal="center"
)

    title.font = Font(
        bold=True,
        size=14
    )

    title.alignment = Alignment(
        horizontal="center"
    )

    # Headers
    ws["A4"] = "Period"

    ws["B4"] = "Class"

    ws["C4"] = "Substitute Teacher"

    header_font = Font(
        bold=True
    )

    ws["A4"].font = header_font

    ws["B4"].font = header_font

    ws["C4"].font = header_font

    row_num = 5

    for row in records:

        teacher = row[2]

        if teacher is None:

            teacher = (
                "Substitute Not Available"
            )

        ws.cell(
            row=row_num,
            column=1,
            value=row[0]
        )

        ws.cell(
            row=row_num,
            column=2,
            value=row[1]
        )

        ws.cell(
            row=row_num,
            column=3,
            value=teacher
        )

        row_num += 1

    # Auto-adjust column width
    for column in ws.iter_cols():

        max_length = 0

        column_number = column[0].column

        for cell in column:

            if cell.value:

                max_length = max(
                    max_length,
                    len(str(cell.value)))

        ws.column_dimensions[
        get_column_letter(
            column_number
        )
    ].width = max_length + 2

    buffer = BytesIO()

    wb.save(
        buffer
    )

    buffer.seek(
        0
    )

    return send_file(

        buffer,

        as_attachment=True,

        download_name=
        f"Substitution_Timetable_{date}.xlsx",

        mimetype=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/export_absences_pdf")
def export_absences_pdf():

    records = get_all_absences_report()

    buffer = BytesIO()

    pdf = SimpleDocTemplate(
    buffer
)

    data = [

        [
            "Teacher",
            "Date",
            "Day"
        ]

    ]

    for row in records:

        data.append(

            [
                row[0],
                str(row[1]),
                row[2]
            ]

        )

    table = Table(
        data
    )

    table.setStyle(

        TableStyle(

            [

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.grey
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.whitesmoke
                ),

                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, -1),
                    colors.beige
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    1,
                    colors.black
                )

            ]

        )

    )

    pdf.build(

        [

            table

        ]

    )
    buffer.seek(0)

    return send_file(
    buffer,
    as_attachment=True,
    download_name="absences_report.pdf",
    mimetype="application/pdf"
)

@app.route(
    "/change_password",
    methods=["GET", "POST"]
)
def change_password():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if request.method == "POST":

        current_password = request.form[
            "current_password"
        ]

        new_password = request.form[
            "new_password"
        ]

        confirm_password = request.form[
            "confirm_password"
        ]

        user = login_user(

            session["username"],

            current_password
        )

        if user is None:

            flash(
                "Current password is incorrect!",
                "danger"
            )

        elif new_password != confirm_password:

            flash(
                "Passwords do not match!",
                "danger"
            )

        else:

            update_user_password(

                session["user_id"],

                new_password
            )

            flash(
                "Password changed successfully!",
                "success"
            )

            if session["role"] == "admin":

                return redirect(
                    url_for(
                        "home"
                    )
                )

            else:

                return redirect(
                    url_for(
                        "teacher_dashboard"
                    )
                )

    return render_template(
        "change_password.html"
    )

@app.route(
    "/change_staff_password",
    methods=["GET", "POST"]
)
def change_staff_password():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if session["role"] != "admin":

        return redirect(
            url_for(
                "teacher_dashboard"
            )
        )

    if request.method == "POST":

        current_password = request.form["current_password"]

        new_password = request.form["new_password"]

        confirm_password = request.form["confirm_password"]

        if current_password != get_staff_password():

            flash(
                "Current staff password is incorrect!",
                "danger"
            )

        elif new_password != confirm_password:

            flash(
                "Passwords do not match!",
                "danger"
            )

        else:

            update_staff_password(
                new_password
            )

            flash(
                "Staff password changed successfully!",
                "success"
            )

            return redirect(
                url_for(
                    "home"
                )
            )

    return render_template(
        "change_password.html"
    )

@app.route("/upcoming_absences")
def upcoming_absences():

    name = request.args.get(
        "name",
        ""
    )

    date = request.args.get(
        "date",
        ""
    )

    if name or date:

        absences = search_upcoming_absences(
            name,
            date
        )

    else:

        absences = fetch_upcoming_absences()

    return render_template(
        "upcoming_absences.html",
        absences=absences,
        search_name=name,
        search_date=date
    )

@app.route("/previous_absences")
def previous_absences():

    name = request.args.get(
        "name",
        ""
    )

    date = request.args.get(
        "date",
        ""
    )

    if name or date:

        absences = search_previous_absences(
            name,
            date
        )

    else:

        absences = fetch_absences()

    return render_template(

        "previous_absences.html",

        absences=absences,

        search_name=name,

        search_date=date
    )

@app.route("/today_substitutions")
def today_substitutions():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    today = datetime.today().strftime(
        "%Y-%m-%d"
    )

    substitutions = get_today_substitutions(
        today
    )

    return render_template(

        "today_substitutions.html",

        substitutions=substitutions,

        today=today
    )

@app.route(
    "/staff_password",
    methods=["GET", "POST"]
)
def staff_password():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if request.method == "POST":

        password = request.form["password"]

        if password == get_staff_password():

            return redirect(
        url_for(
            "today_substitutions"
        )
    )

        flash(
    "Incorrect Staff Password!",
    "danger"
)

    return render_template(
        "staff_password.html"
    )

@app.route("/my_absences")
def my_absences():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if session["role"] != "teacher":

        return redirect(
            url_for(
                "home"
            )
        )

    absences = get_my_absences(
        session["teacher_id"]
    )

    return render_template(

        "my_absences.html",

        absences=absences
    )

@app.route("/my_substitutions")
def my_substitutions():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if session["role"] != "teacher":

        return redirect(
            url_for(
                "home"
            )
        )

    substitutions = get_my_substitutions(
        session["teacher_id"]
    )

    return render_template(

        "my_substitutions.html",

        substitutions=substitutions
    )

@app.route("/my_timetable")
def my_timetable():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if session["role"] != "teacher":

        return redirect(
            url_for(
                "home"
            )
        )

    timetable = get_teacher_full_timetable(
        session["teacher_id"]
    )

    return render_template(

        "my_timetable.html",

        timetable=timetable
    )

@app.route("/teacher_dashboard")
def teacher_dashboard():

    if "role" not in session:

        return redirect(
            url_for(
                "login_page"
            )
        )

    if session["role"] != "teacher":

        return redirect(
            url_for(
                "home"
            )
        )

    teacher = get_teacher_details(
        session["teacher_id"]
    )

    return render_template(

        "teacher_dashboard.html",

        teacher=teacher
    )

@app.route("/previous_substitutions")
def previous_substitutions():

    teacher_name = request.args.get(
        "teacher_name",
        ""
    )

    date = request.args.get(
        "date",
        ""
    )

    if teacher_name or date:

        substitutions = search_previous_substitutions(
            teacher_name,
            date
        )

    else:

        substitutions = fetch_substitutions()

    total_substitutions = len(
        substitutions
    )

    return render_template(
        "previous_substitutions.html",

        substitutions=substitutions,

        search_teacher=teacher_name,

        selected_date=date,

        total_substitutions=total_substitutions
    )

@app.route("/upcoming_substitutions")
def upcoming_substitutions():

    teacher_name = request.args.get(
        "teacher_name",
        ""
    )

    date = request.args.get(
        "date",
        ""
    )

    if teacher_name or date:

        substitutions = search_upcoming_substitutions(
            teacher_name,
            date
        )

    else:

        substitutions = fetch_upcoming_substitutions()

    total_substitutions = len(
        substitutions
    )

    return render_template(
        "upcoming_substitutions.html",

        substitutions=substitutions,

        search_teacher=teacher_name,

        selected_date=date,

        total_substitutions=total_substitutions
    )

@app.route("/substitutions")
def substitutions_page():

    teacher_name = request.args.get(
        "teacher_name",
        ""
    )

    date = request.args.get(
        "date",
        ""
    )

    # If searching, search all records
    if teacher_name or date:

        substitutions = (
            search_today_substitutions_with_absent_teacher(
                teacher_name,
                date
            )
        )

    # Otherwise show today's substitutions
    else:

        today = datetime.today().strftime(
            "%Y-%m-%d"
        )

        substitutions = (
            search_today_substitutions_with_absent_teacher(
                "",
                today
            )
        )

    total_substitutions = len(
        substitutions
    )

    return render_template(

        "substitutions.html",

        substitutions=substitutions,

        search_teacher=teacher_name,

        selected_date=date,

        total_substitutions=total_substitutions
    )

@app.route("/delete_absence/<int:absence_id>")
def delete_absence_page(absence_id):

    delete_absence(absence_id)

    flash(
        "Absence deleted successfully!",
        "danger"
    )

    return redirect(
        url_for(
            "absences_page"
        )
    )

@app.route("/record_absence", methods=["GET", "POST"])
def record_absence_page():

    if request.method == "POST":

        date = request.form["date"]

        day = datetime.strptime(
            date,
            "%Y-%m-%d"
        ).strftime(
            "%A"
        )

        teacher_names = request.form.getlist(
            "teachers"
        )

        if not teacher_names:

            flash(
                "Please select at least one teacher!",
                "danger"
            )

            return redirect(
                url_for(
                    "record_absence_page"
                )
            )

        absence_ids = []

        duplicates = 0

        for teacher_name in teacher_names:

            teacher_id = get_teacher_id(
                teacher_name
            )

            if teacher_id is None:
                flash(f"Teacher '{teacher_name}' not found in the system.", "danger")
                continue

            if absence_exists(
                    teacher_id,
                    date):

                duplicates += 1

                continue

            absence_id = record_absence(
                teacher_id,
                date,
                day
            )

            absence_ids.append(
                absence_id
            )

        # NEW:
        # Store absence IDs temporarily
        # for the suggestion page
        session["pending_absence_ids"] = absence_ids

        if absence_ids:

            count = len(
                absence_ids
            )

            if count == 1:

                flash(
                    "1 absence recorded successfully!",
                    "success"
                )

            else:

                flash(
                    f"{count} absences recorded successfully!",
                    "success"
                )

        if duplicates:

            if duplicates == 1:

                flash(
                    "1 teacher was already marked absent.",
                    "warning"
                )

            else:

                flash(
                    f"{duplicates} teachers were already marked absent.",
                    "warning"
                )

        # NEW:
        # Go to suggestion page
        return redirect(
            url_for(
                "suggest_substitutions"
            )
        )

    teachers = get_teacher_names()

    return render_template(
        "record_absence.html",
        teachers=teachers
    )

@app.route("/absences")
def absences_page():

    name = request.args.get(
        "name",
        ""
    )

    date = request.args.get(
        "date",
        ""
    )

    # If user searches by name or date
    if name or date:

        absences = search_absences(
            name,
            date
        )

    # Otherwise show today's absences
    else:

        today = datetime.today().strftime(
            "%Y-%m-%d"
        )

        absences = get_absences_by_date(
            today
        )

    return render_template(

        "absences.html",

        absences=absences,

        search_name=name,

        search_date=date
    )

@app.route("/add_teacher", methods=["GET", "POST"])
def add_teacher_page():

    classes = [
    "XII-A", "XII-B",
    "XI-A", "XI-B",
    "X-A", "X-B", "X-C",
    "IX-A", "IX-B", "IX-C",
    "VIII-A", "VIII-B", "VIII-C", "VIII-D",
    "VII-A", "VII-B", "VII-C", "VII-D",
    "VI-A", "VI-B", "VI-C", "VI-D",
    "V-A", "V-B", "V-C", "V-D",
    "IV-A", "IV-B", "IV-C", "IV-D",
    "III-A", "III-B", "III-C", "III-D",
    "II-A", "II-B", "II-C", "II-D",
    "I-A", "I-B", "I-C", "I-D",
    "UKG-A", "UKG-B", "UKG-C",
    "LKG-A", "LKG-B", "LKG-C",
    "Pre-KG"]

    if request.method == "POST":

        name = request.form["name"]

        subject = request.form["subject"]

        contact = request.form["contact"]

        classes_selected = request.form.getlist("classes")

        timetable = []

        days = [

            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday"

        ]

        for day in days:

            periods = []

            for i in range(1, 10):

                periods.append(

                    request.form[
                        f"{day}_period{i}"
                    ]

                )

            timetable.append([day] + periods)

        try:
            add_teacher(
                name,
                subject,
                contact,
                timetable,
                classes_selected
            )

            flash(
                "Teacher added successfully!",
                "success"
            )

            return redirect(url_for("add_teacher_page"))

        except ValueError as e:

            flash(str(e), "danger")

            return redirect(url_for("add_teacher_page"))
        
        except IntegrityError:

            flash(f"Teacher with name '{name}' already exists in the system.", "danger")

            return redirect(url_for("add_teacher_page"))

    num_columns = 4
    rows_per_column = math.ceil(len(classes) / num_columns)
    class_columns = [
    classes[i:i + rows_per_column]
    for i in range(0, len(classes), rows_per_column)]

    return render_template(

        "add_teacher.html",

        total_teachers=count_teachers(),

        class_columns=class_columns)

@app.route("/count_absent/<date>")
def count_absent(date):

    count = count_absent_teachers(date)

    return {
        "count": count
    }

@app.route("/export_timetable_pdf/<int:teacher_id>")
def export_timetable_pdf(teacher_id):

    teacher = get_teacher(teacher_id)

    timetable = get_teacher_timetable(
        teacher_id
    )

    if not teacher:

        flash(
            "Teacher not found!",
            "danger"
        )

        return redirect(
            url_for("teachers_page")
        )


    teacher_name = teacher[1]


    # Create PDF in memory

    buffer = BytesIO()


    document = SimpleDocTemplate(

        buffer,

        pagesize=landscape(A4),

        rightMargin=30,

        leftMargin=30,

        topMargin=30,

        bottomMargin=30

    )


    elements = []


    styles = getSampleStyleSheet()


    # School title

    school_title = Paragraph(

        "<b>ALPHA SCHOOL</b>",

        styles["Title"]

    )


    elements.append(
        school_title
    )


    elements.append(
        Spacer(1, 10)
    )


    # Timetable title

    timetable_title = Paragraph(

        "<b>TEACHER TIMETABLE</b>",

        styles["Heading2"]

    )


    elements.append(
        timetable_title
    )


    elements.append(
        Spacer(1, 10)
    )


    # Teacher name

    teacher_paragraph = Paragraph(

        f"<b>Teacher:</b> {teacher_name}",

        styles["Heading3"]

    )


    elements.append(
        teacher_paragraph
    )


    elements.append(
        Spacer(1, 20)
    )


    # Table headings

    table_data = [

        [
            "Day",
            "P1",
            "P2",
            "P3",
            "P4",
            "P5",
            "P6",
            "P7",
            "P8",
            "P9"
        ]

    ]


    # Add timetable rows

    for row in timetable:

        table_data.append(
            list(row)
        )


    table = Table(

        table_data,

        repeatRows=1

    )


    table.setStyle(

        TableStyle(

            [

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.lightgrey
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.black
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    1,
                    colors.black
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                )

            ]

        )

    )


    elements.append(
        table
    )


    document.build(
        elements
    )


    buffer.seek(0)


    # Safe filename

    safe_teacher_name = (
        teacher_name
        .replace(" ", "_")
        .replace("/", "_")
    )


    return send_file(

        buffer,

        as_attachment=True,

        download_name=
        f"{safe_teacher_name}_Timetable.pdf",

        mimetype="application/pdf"

    )

@app.route("/export_timetable_excel/<int:teacher_id>")
def export_timetable_excel(teacher_id):

    teacher = get_teacher(teacher_id)

    timetable = get_teacher_timetable(
        teacher_id
    )


    if not teacher:

        flash(
            "Teacher not found!",
            "danger"
        )

        return redirect(
            url_for("teachers_page")
        )


    teacher_name = teacher[1]


    # Create workbook

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Timetable"


    # Main title

    worksheet.merge_cells(
        "A1:J1"
    )


    worksheet["A1"] = (
        "ALPHA SCHOOL"
    )


    worksheet["A1"].font = Font(

        bold=True,

        size=18

    )


    worksheet["A1"].alignment = Alignment(

        horizontal="center"

    )


    # Timetable heading

    worksheet.merge_cells(
        "A2:J2"
    )


    worksheet["A2"] = (
        "TEACHER TIMETABLE"
    )


    worksheet["A2"].font = Font(

        bold=True,

        size=14

    )


    worksheet["A2"].alignment = Alignment(

        horizontal="center"

    )


    # Teacher name

    worksheet.merge_cells(
        "A3:J3"
    )


    worksheet["A3"] = (
        f"Teacher: {teacher_name}"
    )


    worksheet["A3"].font = Font(

        bold=True

    )


    worksheet["A3"].alignment = Alignment(

        horizontal="center"

    )


    # Headers

    headers = [

        "Day",

        "P1",

        "P2",

        "P3",

        "P4",

        "P5",

        "P6",

        "P7",

        "P8",

        "P9"

    ]


    for column, heading in enumerate(

        headers,

        start=1

    ):

        cell = worksheet.cell(

            row=5,

            column=column,

            value=heading

        )


        cell.font = Font(

            bold=True

        )


        cell.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )


    # Add timetable data

    for row_number, row in enumerate(

        timetable,

        start=6

    ):

        for column_number, value in enumerate(

            row,

            start=1

        ):

            cell = worksheet.cell(

                row=row_number,

                column=column_number,

                value=value

            )


            cell.alignment = Alignment(

                horizontal="center",

                vertical="center"

            )


    # Borders

    thin_border = Border(

        left=Side(style="thin"),

        right=Side(style="thin"),

        top=Side(style="thin"),

        bottom=Side(style="thin")

    )


    for row in worksheet.iter_rows(

        min_row=5,

        max_row=worksheet.max_row,

        min_col=1,

        max_col=10

    ):

        for cell in row:

            cell.border = thin_border


    # Automatic column widths

    for column in range(

        1,

        11

    ):

        max_length = 0


        column_letter = get_column_letter(

            column

        )


        for cell in worksheet[

            column_letter

        ]:

            if cell.value:

                max_length = max(

                    max_length,

                    len(str(cell.value))

                )


        worksheet.column_dimensions[

            column_letter

        ].width = (

            max_length + 4

        )


    # Save Excel in memory

    buffer = BytesIO()


    workbook.save(
        buffer
    )


    buffer.seek(0)


    safe_teacher_name = (
        teacher_name
        .replace(" ", "_")
        .replace("/", "_")
    )


    return send_file(

        buffer,

        as_attachment=True,

        download_name=
        f"{safe_teacher_name}_Timetable.xlsx",

        mimetype=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

@app.route("/timetable/<int:teacher_id>")
def timetable_page(teacher_id):

    timetable = get_teacher_timetable(
        teacher_id
    )

    return render_template(
        "timetable.html",
        timetable=timetable,
        teacher_id=teacher_id
    )

@app.route("/edit_teacher/<int:teacher_id>", methods=["GET", "POST"])
def edit_teacher_page(teacher_id):

    if request.method == "POST":

        name = request.form["name"]
        subject = request.form["subject"]
        contact = request.form["contact"]

        selected_classes = request.form.getlist(
            "classes"
        )

        # Prepare timetable
        timetable = []

        days = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday"
        ]

        for day in days:

            periods = []

            for i in range(1, 10):

                periods.append(
                    request.form[
                        f"{day}_period{i}"
                    ]
                )

            timetable.append(
                [day] + periods
            )

        # Update teacher information,
        # capable classes and timetable
        # using ONE database connection

        update_teacher_full(
            teacher_id,
            name,
            subject,
            contact,
            selected_classes,
            timetable
        )

        flash(
            "Changes saved successfully!",
            "success"
        )

        return redirect(
            url_for(
                "teachers_page"
            )
        )


    # GET request

    teacher = get_teacher(
        teacher_id
    )

    timetable = get_full_timetable(
        teacher_id
    )

    teacher_classes = get_teacher_classes(
        teacher_id
    )

    classes = [
        "XII-A", "XII-B",
        "XI-A", "XI-B",
        "X-A", "X-B", "X-C",
        "IX-A", "IX-B", "IX-C",
        "VIII-A", "VIII-B", "VIII-C", "VIII-D",
        "VII-A", "VII-B", "VII-C", "VII-D",
        "VI-A", "VI-B", "VI-C", "VI-D",
        "V-A", "V-B", "V-C", "V-D",
        "IV-A", "IV-B", "IV-C", "IV-D",
        "III-A", "III-B", "III-C", "III-D",
        "II-A", "II-B", "II-C", "II-D",
        "I-A", "I-B", "I-C", "I-D",
        "UKG-A", "UKG-B", "UKG-C",
        "LKG-A", "LKG-B", "LKG-C",
        "Pre-KG"
    ]

    num_columns = 4

    rows_per_column = math.ceil(
        len(classes) / num_columns
    )

    class_columns = [

        classes[
            i:i + rows_per_column
        ]

        for i in range(
            0,
            len(classes),
            rows_per_column
        )
    ]

    return render_template(
        "edit_teacher.html",
        teacher=teacher,
        timetable=timetable,
        teacher_classes=teacher_classes,
        class_columns=class_columns
    )

@app.route("/delete_teacher/<int:teacher_id>")
def delete_teacher_page(teacher_id):

    delete_teacher(teacher_id)

    flash(
    "Teacher deleted successfully!",
    "danger"
)

    return redirect(url_for("teachers_page"))

@app.route("/teachers")
def teachers_page():

    name = request.args.get(
        "name",
        ""
    )

    if name:

        teachers = search_teachers(
            name
        )

    else:

        teachers = fetch_all_teachers()

    return render_template(
        "teachers.html",

        teachers=teachers,

        search_name=name
    )


if __name__ == "__main__":
    app.run(debug=True)